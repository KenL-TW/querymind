"""
Export endpoints: execute SQL and download results as CSV or XLSX.

GET  /v1/export/csv?sql=...&conn_name=...&filename=...
GET  /v1/export/xlsx?sql=...&conn_name=...&filename=...
POST /v1/export/csv    (body: ExportRequest — for long SQL bodies)
POST /v1/export/xlsx   (body: ExportRequest)

Requires authenticated user with ``role.can_export == True``.
Enforces RBAC SQL allow-list, applies row cap, and guards against CSV
formula injection (``=`` / ``+`` / ``-`` / ``@`` prefixes).
"""
from __future__ import annotations

import csv
import io
import logging
import time

from fastapi import APIRouter, Body, Depends, HTTPException, Query, Request
from fastapi.responses import Response

from api.auth import require_user
from api.rate_limit import limiter
from api.schemas import ExportRequest
from config.settings import settings
from core.dlp import mask_rows_with_report
from core.rbac import (
    PermissionDeniedError,
    UserContext,
    assert_conn_allowed,
    assert_sql_allowed,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/v1/export", tags=["export"])


# ─── helpers ───────────────────────────────────────────────────────────────

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _csv_safe(value) -> str:
    """Escape spreadsheet formula-injection vectors per OWASP guidance.

    Prefixing with a single quote forces Excel/LibreOffice to treat the cell
    as text and not evaluate it.  Only applied to string values.
    """
    if value is None:
        return ""
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return str(value)


def _safe_filename(name: str) -> str:
    cleaned = "".join(c for c in name if c.isalnum() or c in "-_")[:64]
    return cleaned or "export"


def _audit(audit_logger, *, event: str, user: UserContext, conn_name: str,
           sql: str, duration_ms: int, status: str, error: str | None = None,
           dlp_report: dict | None = None) -> None:
    if not audit_logger:
        return
    dlp_summary = ""
    if dlp_report:
        cols = ",".join(str(c.get("column")) for c in dlp_report.get("columns", [])[:20])
        dlp_summary = (
            f" dlp_applied={bool(dlp_report.get('applied'))}"
            f" dlp_redactions={int(dlp_report.get('total_redactions') or 0)}"
            f" dlp_columns={cols}"
        )
    detail = f"role={user.role.name}{dlp_summary} sql={sql[:6000]}"
    audit_logger.log(
        event,
        session_id=None,
        api_key_prefix=(user.api_key_prefix or user.user_id or "anon")[:8],
        conn_name=conn_name,
        status=status,
        duration_ms=duration_ms,
        detail=detail,
        error_msg=(error or "")[:8000] or None,
    )


def _run_export_query(user: UserContext, sql: str, conn_name: str) -> tuple[list[dict], dict]:
    """Common path used by GET + POST handlers.

    Performs RBAC checks, executes the query through the standard connector,
    applies the role's row cap as a hard upper bound (defense-in-depth in case
    the SQL omits LIMIT), and returns the rows as dicts.
    """
    if not user.role.can_export:
        raise HTTPException(status_code=403, detail=f"角色 {user.role.name} 沒有匯出權限。")

    try:
        assert_conn_allowed(user, conn_name)
        verb = assert_sql_allowed(user, sql)
    except PermissionDeniedError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    if verb not in {"SELECT", "WITH"}:
        raise HTTPException(status_code=403, detail="匯出僅允許 SELECT/WITH 查詢。")

    from api.main import app_state
    from tools.db_tools import _apply_row_cap

    registry = app_state["registry"]
    try:
        conn = registry.get(conn_name)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    try:
        capped_sql = _apply_row_cap(sql, user.role.max_rows_per_query)
        rows = conn.execute(capped_sql)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if isinstance(rows, list) and len(rows) > user.role.max_rows_per_query:
        rows = rows[: user.role.max_rows_per_query]
    rows = rows or []
    exempt = {r.strip() for r in (settings.dlp_role_exempt or "").split(",") if r.strip()}
    masked_rows, dlp_report = mask_rows_with_report(
        rows,
        enabled=bool(settings.dlp_enabled),
        role_exempt=exempt,
        role_name=getattr(user, "role_name", None),
    )
    return masked_rows, dlp_report


def _dlp_headers(dlp_report: dict) -> dict[str, str]:
    columns = ",".join(str(c.get("column")) for c in dlp_report.get("columns", [])[:20])
    return {
        "X-DLP-Enabled": "true" if dlp_report.get("enabled") else "false",
        "X-DLP-Applied": "true" if dlp_report.get("applied") else "false",
        "X-DLP-Redactions": str(int(dlp_report.get("total_redactions") or 0)),
        "X-DLP-Columns": columns,
    }


def _build_csv(rows: list[dict]) -> bytes:
    if not rows:
        return "\ufeff".encode("utf-8")
    output = io.StringIO()
    fieldnames = list(rows[0].keys())
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow({k: _csv_safe(row.get(k)) for k in fieldnames})
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _build_xlsx(rows: list[dict], sheet_name: str) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Sheet1")[:31]

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            ws.append([
                _csv_safe(row.get(h)) if (
                    isinstance(row.get(h), str)
                    and row.get(h)
                    and row.get(h)[0] in _FORMULA_PREFIXES
                ) else row.get(h)
                for h in headers
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ─── GET handlers (short queries via URL) ─────────────────────────────────

@router.get("/csv")
@limiter.limit(settings.rate_limit_api)
async def export_csv(
    request: Request,
    sql: str = Query(..., description="SQL SELECT statement"),
    conn_name: str = Query(default="default"),
    filename: str = Query(default="export"),
    user: UserContext = Depends(require_user),
) -> Response:
    return _export_csv_response(sql, conn_name, filename, user)


@router.get("/xlsx")
@limiter.limit(settings.rate_limit_api)
async def export_xlsx(
    request: Request,
    sql: str = Query(..., description="SQL SELECT statement"),
    conn_name: str = Query(default="default"),
    filename: str = Query(default="export"),
    sheet_name: str = Query(default="Sheet1"),
    user: UserContext = Depends(require_user),
) -> Response:
    return _export_xlsx_response(sql, conn_name, filename, sheet_name, user)


# ─── POST handlers (long queries via body — avoid URL length limits) ──────

@router.post("/csv")
@limiter.limit(settings.rate_limit_api)
async def export_csv_post(
    request: Request,
    body: ExportRequest = Body(...),
    user: UserContext = Depends(require_user),
) -> Response:
    return _export_csv_response(body.sql, body.conn_name, body.filename, user)


@router.post("/xlsx")
@limiter.limit(settings.rate_limit_api)
async def export_xlsx_post(
    request: Request,
    body: ExportRequest = Body(...),
    user: UserContext = Depends(require_user),
) -> Response:
    return _export_xlsx_response(
        body.sql, body.conn_name, body.filename, body.sheet_name, user
    )


# ─── shared response builders ──────────────────────────────────────────────

def _export_csv_response(
    sql: str, conn_name: str, filename: str, user: UserContext,
) -> Response:
    from api.main import app_state
    audit_logger = app_state.get("audit_logger")
    t_start = time.monotonic()
    try:
        rows, dlp_report = _run_export_query(user, sql, conn_name)
    except HTTPException as exc:
        _audit(audit_logger, event="export_csv", user=user, conn_name=conn_name,
               sql=sql, duration_ms=int((time.monotonic() - t_start) * 1000),
               status="error", error=str(exc.detail))
        raise
    content = _build_csv(rows)
    duration_ms = int((time.monotonic() - t_start) * 1000)
    _audit(audit_logger, event="export_csv", user=user, conn_name=conn_name,
           sql=sql, duration_ms=duration_ms, status="success", dlp_report=dlp_report)
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{_safe_filename(filename)}.csv"',
            "X-Row-Count": str(len(rows)),
            **_dlp_headers(dlp_report),
        },
    )


def _export_xlsx_response(
    sql: str, conn_name: str, filename: str, sheet_name: str, user: UserContext,
) -> Response:
    from api.main import app_state
    audit_logger = app_state.get("audit_logger")
    t_start = time.monotonic()
    try:
        rows, dlp_report = _run_export_query(user, sql, conn_name)
    except HTTPException as exc:
        _audit(audit_logger, event="export_xlsx", user=user, conn_name=conn_name,
               sql=sql, duration_ms=int((time.monotonic() - t_start) * 1000),
               status="error", error=str(exc.detail))
        raise
    content = _build_xlsx(rows, sheet_name)
    duration_ms = int((time.monotonic() - t_start) * 1000)
    _audit(audit_logger, event="export_xlsx", user=user, conn_name=conn_name,
           sql=sql, duration_ms=duration_ms, status="success", dlp_report=dlp_report)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{_safe_filename(filename)}.xlsx"',
            "X-Row-Count": str(len(rows)),
            **_dlp_headers(dlp_report),
        },
    )
